#!/usr/bin/env python3
"""
Add the 25 stations that are in the June-12 complete lists but missing from
stations.json (they were announced after the 2026-06-05 PSMS extract).

Name/city/domain come from the Student Preference export (reliable .xlsx);
stipend/tags/project-domain were read from the June-12 announced + project-title
sheets. Visited flags are recomputed over ALL stations from the PS Tracker .xlsx.

Run:  python3 add_missing.py --write     (updates stations.json + re-embeds into index.html)
"""
import json, re, os, sys, datetime
from pathlib import Path
import openpyxl

HERE = Path(__file__).resolve().parent
DL = Path(os.path.expanduser("~/Downloads"))
STATIONS_JSON = HERE / "stations.json"
TRACK = DL / "PS Tracker - Sem 1 2026-27.xlsx"
WRITE = "--write" in sys.argv

CODE_MAP = {
    'A1':'Chemical Engineering','A2':'Civil Engineering','A3':'EEE',
    'A4':'Mechanical Engineering','A5':'B.Pharm','A7':'CSE',
    'A8':'Electronics & Instrumentation','AA':'ECE','AB':'Manufacturing Engineering',
    'AC':'Electronics & Computer Engineering','AD':'Mathematics and Computing',
    'AJ':'Environmental & Sustainability Engineering','B1':'Biological Sciences',
    'B2':'Chemistry','B3':'Economics','B4':'Mathematics','B5':'Physics',
    'B7':'Semiconductor & Nanoscience',
}
def br(*codes):
    return [CODE_MAP[c] for c in codes if c in CODE_MAP]

CITY_STATE = {
    "Hyderabad":"Telangana","Bengaluru":"Karnataka","Ahmedabad":"Gujarat",
    "Mumbai":"Maharashtra","Gurgaon":"Haryana","New Delhi":"Delhi","Hosur":"Tamil Nadu",
}

# id -> (stipend, projectTitle, projectDomains, branches)
EXTRA = {
    6247: (0,      "Research — multiple projects (lasers, NMR, biophysics, theory)",
                   ["Physics","Data Science","Biology","Chemistry","Mechanical","Computer Science"],
                   br('B5','A3','A8','A7','B4','A1','B2','A4','B1')),
    6546: (35000,  "Designing Algorithm for HFT (High Frequency Trading)",
                   ["FinTech"], br('A8','A3','A7','B3','B1','B2','B4','B5','AA')),
    6559: (20000,  "Standard Operating Manual / Drone Delivery / e-Medic Funnel", [], []),
    7519: (30000,  "AIML Intern — Requirement Engineering & Use Case Development",
                   ["AI / ML"], br('A3','A7','AA')),
    7865: (40000,  "Graduate Engineer Trainees", [], br('A7')),
    8410: (0,      "Intern", [], []),
    9253: (60000,  "Frontier systems-design", [], br('B2','AB')),
    9272: (30000,  "Cell Engineering — details awaited", [], br('A1')),
    9273: (30000,  "Chemical Engineering — details awaited", [], br('A1')),
    9274: (15000,  "Intern", [], []),
    9277: (30000,  "Growth Marketing — details awaited", ["Marketing"], []),
    9278: (40000,  "Data Science — details awaited", ["Data Science"], []),
    9281: (60000,  "Healthcare Product — details awaited", ["Product Management"],
                   br('A1','A2','A3','A4','A5','A7','A8','AA','AB','AD','AJ','B1','B2','B3','B4','B5')),
    9282: (0, "Flexible Pavements Division — details awaited", [], br('A2','A7','A8')),
    9283: (0, "Rigid Pavements Division — details awaited", [], br('A2','A7')),
    9284: (0, "Geotechnical Engineering Division — details awaited", [], br('A2','A7')),
    9285: (0, "Bridge Engineering & Structures Division — details awaited", [], br('A2','A8')),
    9286: (0, "Traffic Engineering & Safety Division — details awaited", [], br('A2','A7')),
    9287: (0, "Transport Planning & Environment Division — details awaited", [], br('A2','A7')),
    9288: (0, "Pavement Evaluation Division — details awaited", [], br('A2','A8')),
    9289: (0, "Planning, Monitoring & Evaluation (PME) — details awaited", [], br('A2','A7')),
    9290: (0, "Computer Center & Networking — details awaited", [], br('A2','A7')),
    9291: (0, "Information, Liaison & Training (ILT) — details awaited", [], br('A2','A7')),
    9292: (0, "Knowledge Resource Centre (KRC) — details awaited", [], br('A2','A7')),
    9293: (0, "Engineering Service Division (ESD) — details awaited", [], br('A2','A7')),
}

# ---- name/city/domain from the preference export ----
wb = openpyxl.load_workbook(DL / "Student_Preference_Station_export_07062026083410.xlsx",
                            read_only=True, data_only=True)
ws = wb.active; rows = ws.iter_rows(values_only=True)
hdr = [str(c).strip() for c in next(rows)]; ix = {h: i for i, h in enumerate(hdr)}
pref = {}
for r in rows:
    try: sid = int(r[ix['Station ID']])
    except (TypeError, ValueError): continue
    pref.setdefault(sid, {'name': r[ix['Station Name']], 'city': r[ix['Centre (City)']],
                          'domain': r[ix['Station Domain']]})
wb.close()

# ---- interaction companies from the PS Tracker ----
def clean(s): return ("" if s is None else str(s)).replace("\xa0", " ").strip()
wb = openpyxl.load_workbook(TRACK, read_only=True, data_only=True)
ws = wb["PS Tracker"] if "PS Tracker" in wb.sheetnames else wb.active
rows = ws.iter_rows(values_only=True); thdr = [clean(c) for c in next(rows)]
ci = thdr.index("Company")
companies = []
for r in rows:
    c = clean(r[ci]) if ci < len(r) else ""
    if c and c.lower() != "company" and c not in companies:
        companies.append(c)
wb.close()

def normalize(name):
    n = re.sub(r"[^a-z0-9 ]", " ", name.lower())
    n = re.sub(r"\s+", " ", n).strip()
    for w in (" limited"," ltd"," pvt"," private"," inc"," llp"," co "," india"," technologies",
              " technology"," services"," solutions"," systems"," group"," labs"," lab"," corp"," company"):
        n = n.replace(w, " ")
    return re.sub(r"\s+", " ", n).strip()
def nospace(s): return s.replace(" ", "")
def matches(company, station):
    cn, sn = normalize(company), normalize(station); cl = company.lower()
    if not cn or not sn: return False
    if cl.startswith("csir-4pi") or cl == "csir-4pi": return station.strip().lower() == "csir-4pi"
    if "siemens eda" in cl: s=station.lower(); return "siemens" in s and "eda" in s
    if "siemens energy" in cl: return "siemens energy" in station.lower()
    if cl.startswith("itc"): return station.lower().startswith("itc")
    if cl.startswith("v-gaurd") or cl.startswith("v-guard"): return station.lower().startswith("v-guard")
    if "qapita" in cl: return "qapita" in station.lower()
    strict = (cn == sn) or sn.startswith(cn+" ") or cn.startswith(sn+" ")
    if not strict and len(nospace(cn)) >= 6 and len(nospace(sn)) >= 6:
        strict = nospace(sn).startswith(nospace(cn)) or nospace(cn).startswith(nospace(sn))
    return strict

# the on-disk tracker is the May-28 snapshot; add interaction companies announced
# after it that are relevant to the new stations (existing flags came from a newer tracker)
companies_new = companies + ["KPMG", "Pfizer"]

# ---- build + append (existing 554 stations' visited flags are left untouched) ----
doc = json.load(open(STATIONS_JSON))
stations = doc["stations"]
have = {int(s["stationId"]) for s in stations if str(s["stationId"]).isdigit()}
added = 0
for sid, (stipend, title, domains, branches) in EXTRA.items():
    if sid in have:
        continue
    p = pref.get(sid, {})
    name = p.get("name") or title
    city = p.get("city") or ""
    dom = p.get("domain") or "Others"
    hits = [c for c in companies_new if matches(c, name)]
    stations.append({
        "stationId": str(sid), "stationName": name, "businessDomain": dom,
        "address": "", "state": CITY_STATE.get(city, city), "country": "India",
        "projectId": "", "projectTitle": title, "description": "",
        "mentorName": "", "mentorEmail": "", "stipend": float(stipend), "currency": "INR",
        "workMode": "", "officeHours": "", "weekHolidays": "",
        "cgpaMin": None, "cgpaMax": None, "intake": None,
        "eligibleBranches": "", "branches": branches, "techStack": [], "stipendPct": 0,
        "projectDomains": domains, "bizDomain": dom, "city": city,
        "visited": bool(hits), "visitedCompanies": hits,
    })
    added += 1

nvis = sum(1 for s in stations if s.get("visited"))
doc["total"] = len(stations)
doc["with_details"] = sum(1 for s in stations if s.get("projectTitle") or s.get("mentorName") or s.get("stipend"))
doc["updated_at"] = datetime.datetime.now().isoformat(timespec="seconds")

print(f"added {added} missing stations  |  total now {len(stations)}  |  visited {nvis}")

def embed(txt):
    idx = HERE / "index.html"
    html = idx.read_text(encoding="utf-8")
    block = ('<script type="application/json" id="dataStore">\n'
             + txt.replace("</", "<\\/") + "\n</script>\n")
    html = re.sub(r'<script type="application/json" id="dataStore">.*?</script>\n?',
                  "", html, flags=re.S)
    i = html.index("\n<script>\n")
    idx.write_text(html[:i+1] + block + html[i+1:], encoding="utf-8")
    print("re-embedded into index.html")

if WRITE:
    txt = json.dumps(doc, ensure_ascii=False, indent=0)
    STATIONS_JSON.write_text(txt, encoding="utf-8")
    print("wrote stations.json")
    embed(txt)
else:
    print("(dry run — pass --write)")
