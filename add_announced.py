#!/usr/bin/env python3
"""
Reconcile stations.json against the full ANNOUNCED list (June-12) parsed from the
on-disk PDF: add every station still missing, and backfill `city` for all stations
so the UI can show "City, State". Visited flags for existing stations are preserved;
new ones are matched against the PS Tracker (+ post-snapshot KPMG/Pfizer).

Run:  python3 add_announced.py            # dry run (prints sample + counts)
      python3 add_announced.py --write    # update stations.json + re-embed index.html
"""
import os, re, json, sys, datetime
from pathlib import Path
from pypdf import PdfReader
import openpyxl

HERE = Path(__file__).resolve().parent
DL = Path(os.path.expanduser("~/Downloads"))
STATIONS_JSON = HERE / "stations.json"
WRITE = "--write" in sys.argv

ANN_PDF = DL / "List of stations announced - June 12th, 2026.xlsx - data.pdf"
PROJ_PDF = DL / "List of stations with project titles and description - June 12th, 2026.xlsx - data.pdf"
TRACK = DL / "PS Tracker - Sem 1 2026-27.xlsx"

CODE_MAP = {'A1':'Chemical Engineering','A2':'Civil Engineering','A3':'EEE','A4':'Mechanical Engineering',
    'A5':'B.Pharm','A7':'CSE','A8':'Electronics & Instrumentation','AA':'ECE','AB':'Manufacturing Engineering',
    'AC':'Electronics & Computer Engineering','AD':'Mathematics and Computing',
    'AJ':'Environmental & Sustainability Engineering','B1':'Biological Sciences','B2':'Chemistry',
    'B3':'Economics','B4':'Mathematics','B5':'Physics','B7':'Semiconductor & Nanoscience'}
BIZ = ['Finance and Mgmt','Health Care','CSIS/IT','Electronics','Mechanical','Chemical',
       'Manufacturing','Infrastructure','Others']
DATE_RE = re.compile(r'((?:January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{1,2}(?:st|nd|rd|th),\s*20\d\d)')

CITY_STATE = {
    'Bengaluru':'Karnataka','Bangalore':'Karnataka','Manipal':'Karnataka','Sirsi':'Karnataka',
    'Hyderabad':'Telangana','Bhadrachalam':'Telangana','Suriapet':'Telangana',
    'Mumbai':'Maharashtra','Pune':'Maharashtra','Thane':'Maharashtra','Navi Mumbai':'Maharashtra',
    'Nagpur':'Maharashtra','Nagapur':'Maharashtra','Taloja':'Maharashtra','Satara':'Maharashtra',
    'Nagda':'Madhya Pradesh','Indore':'Madhya Pradesh',
    'Gurgaon':'Haryana','Gurugram':'Haryana','Gurgaon-Bawal':'Haryana','Bawal':'Haryana','Ballabgarh':'Haryana',
    'Noida':'Uttar Pradesh','Renukoot':'Uttar Pradesh','Lucknow':'Uttar Pradesh','Kanpur':'Uttar Pradesh',
    'New Delhi':'Delhi','Delhi':'Delhi',
    'Kota':'Rajasthan','Jaipur':'Rajasthan','Pilani':'Rajasthan',
    'Chennai':'Tamil Nadu','Coimbatore':'Tamil Nadu','Hosur':'Tamil Nadu',
    'Ahmedabad':'Gujarat','Gandhinagar':'Gujarat','Vadodara':'Gujarat','Rajkot':'Gujarat','Nadiad':'Gujarat','Dahej':'Gujarat',
    'Kolkata':'West Bengal','Kalyani':'West Bengal','Samali':'West Bengal',
    'Goa':'Goa','Panaji':'Goa','Mohali':'Punjab','Kochi':'Kerala','Thiruvananthapuram':'Kerala',
    'Dehradun':'Uttarakhand','Haridwar':'Uttarakhand','Guwahati':'Assam',
}
CENTRES = sorted(set(CITY_STATE) | {
    'PAN India','Mumbai (Worli)','Indore & Noida','Pune / Hyderabad',
    'Gurgaon / Bengaluru / Mumbai','Pune / Chennai / Bengaluru / Delhi','Gurugram/ Bengaluru/ Chennai',
    'Bengaluru/Mumbai','Bengaluru/Pune','Bengaluru/Kolkata','Hyderabad/Bengaluru','Mumbai/Bengaluru',
    'Mumbai/Hyderabad','Mumbai/Pune','Bengaluru/Mumbai/Hyderabad','GIDC Navsari, Gujarat',
    'Sanga Reddy District Telangana 502319','Uttar Pradesh, Maharashtra & Gujarat',
    'Orlando','California','USA','New York','San Francisco','Dubai','AMBA',
}, key=len, reverse=True)

def pdf_text(p): return "\n".join(pg.extract_text() or "" for pg in PdfReader(str(p)).pages)

# ---- announced: id -> name, city, stipend ----
ann = {}
for ln in pdf_text(ANN_PDF).splitlines():
    m = re.match(r'\s*(\d{4})\s+(.*)$', ln)
    if not m: continue
    sid = int(m.group(1)); rest = m.group(2)
    dm = DATE_RE.search(rest)
    mid = rest[:dm.start()].strip() if dm else rest.strip()
    tail = rest[dm.end():].strip() if dm else ""
    sm = re.search(r'(-?\d[\d,]*)\s*$', tail)
    stipend = float(sm.group(1).replace(',', '')) if sm else 0.0
    # city is always the trailing column; pypdf sometimes glues it to the name
    city, name = "", mid
    for c in CENTRES:                       # longest-first
        if mid == c or mid.endswith(c):
            city = c; name = mid[:len(mid)-len(c)].rstrip(" -,"); break
    if not city and " " in mid:
        city = mid.rsplit(" ", 1)[1]; name = mid.rsplit(" ", 1)[0]
    ann[sid] = {"name": (name or mid).rstrip(" -,"), "city": city, "stipend": stipend}

# ---- project-titles: id -> bizDomain, tags ----
proj = {}
for ln in pdf_text(PROJ_PDF).splitlines():
    m = re.match(r'\s*(\d{4})\s+(.*)$', ln)
    if not m: continue
    sid = int(m.group(1)); rest = m.group(2)
    bd, tags = "", ""
    bm = None
    for d in BIZ:
        mm = re.search(re.escape(d) + r'\s*(\d+)', rest)
        if mm and (bm is None or mm.start() > bm.start()):  # last domain-then-digit wins
            bm, bd = mm, d
    if bm:
        after = rest[bm.end():]
        tm = re.split(r'(?:Description|Title|Details awaited|We are|Job |About |Role|\bThe \b)', after, 1)[0]
        tags = tm.strip(" -,")
    proj.setdefault(sid, {"bizDomain": bd, "tags": tags})

def parse_branches(tagstr):
    out = []
    for tok in str(tagstr).split(','):
        tok = tok.strip()
        if not tok or tok == '-': continue
        core = tok[3:] if tok.startswith('Any') else tok
        chunks = [core[i:i+2] for i in range(0, len(core), 2)]
        if core and len(core) % 2 == 0 and all(c in CODE_MAP for c in chunks):
            for c in chunks:
                if CODE_MAP[c] not in out: out.append(CODE_MAP[c])
    return out

# ---- tracker companies ----
def clean(s): return ("" if s is None else str(s)).replace("\xa0", " ").strip()
wb = openpyxl.load_workbook(TRACK, read_only=True, data_only=True)
ws = wb["PS Tracker"] if "PS Tracker" in wb.sheetnames else wb.active
rows = ws.iter_rows(values_only=True); th = [clean(c) for c in next(rows)]; ci = th.index("Company")
companies = []
for r in rows:
    c = clean(r[ci]) if ci < len(r) else ""
    if c and c.lower() != "company" and c not in companies: companies.append(c)
wb.close()
companies += ["KPMG", "Pfizer", "Blinkit", "Paytm"]

def nz(name):
    n = re.sub(r"[^a-z0-9 ]", " ", name.lower()); n = re.sub(r"\s+", " ", n).strip()
    for w in (" limited"," ltd"," pvt"," private"," inc"," llp"," co "," india"," technologies"," technology",
              " services"," solutions"," systems"," group"," labs"," lab"," corp"," company"):
        n = n.replace(w, " ")
    return re.sub(r"\s+", " ", n).strip()
def nsp(s): return s.replace(" ", "")
def matches(company, station):
    cn, sn = nz(company), nz(station); cl = company.lower()
    if not cn or not sn: return False
    if cl.startswith("csir-4pi") or cl == "csir-4pi": return station.strip().lower() == "csir-4pi"
    if "siemens eda" in cl: s = station.lower(); return "siemens" in s and "eda" in s
    if "siemens energy" in cl: return "siemens energy" in station.lower()
    if cl.startswith("itc"): return station.lower().startswith("itc")
    if cl.startswith("v-gaurd") or cl.startswith("v-guard"): return station.lower().startswith("v-guard")
    if "qapita" in cl: return "qapita" in station.lower()
    strict = (cn == sn) or sn.startswith(cn + " ") or cn.startswith(sn + " ")
    if not strict and len(nsp(cn)) >= 6 and len(nsp(sn)) >= 6:
        strict = nsp(sn).startswith(nsp(cn)) or nsp(cn).startswith(nsp(sn))
    return strict

# ---- reconcile ----
doc = json.load(open(STATIONS_JSON)); stations = doc["stations"]
have = {int(s["stationId"]) for s in stations if str(s["stationId"]).isdigit()}

# backfill city for existing stations
filled = 0
for s in stations:
    try: sid = int(s["stationId"])
    except (ValueError, KeyError): continue
    if not s.get("city") and sid in ann and ann[sid]["city"]:
        s["city"] = ann[sid]["city"]; filled += 1

added = []
for sid in sorted(set(ann) - have):
    a = ann[sid]; p = proj.get(sid, {})
    name = a["name"]; city = a["city"]
    bd = p.get("bizDomain") or "Others"
    branches = parse_branches(p.get("tags", ""))
    hits = [c for c in companies if matches(c, name)]
    stations.append({
        "stationId": str(sid), "stationName": name, "businessDomain": bd,
        "address": "", "state": CITY_STATE.get(city, ""), "country": "India",
        "projectId": "", "projectTitle": "", "description": "",
        "mentorName": "", "mentorEmail": "", "stipend": a["stipend"], "currency": "INR",
        "workMode": "", "officeHours": "", "weekHolidays": "",
        "cgpaMin": None, "cgpaMax": None, "intake": None,
        "eligibleBranches": p.get("tags", ""), "branches": branches, "techStack": [], "stipendPct": 0,
        "projectDomains": [], "bizDomain": bd, "city": city,
        "visited": bool(hits), "visitedCompanies": hits,
    })
    added.append(sid)

doc["total"] = len(stations)
doc["with_details"] = sum(1 for s in stations if s.get("projectTitle") or s.get("mentorName") or s.get("stipend"))
doc["updated_at"] = datetime.datetime.now().isoformat(timespec="seconds")

print(f"announced parsed: {len(ann)} | project parsed: {len(proj)}")
print(f"city backfilled on existing: {filled}")
print(f"added: {len(added)} | total now: {len(stations)} | visited: {sum(1 for s in stations if s['visited'])}")
print("sample added:")
for sid in added[:18]:
    s = next(x for x in stations if x['stationId'] == str(sid))
    print(f"   {sid}  {s['businessDomain']:16} {(s['city']+', '+s['state']).strip(', '):26} {'★' if s['visited'] else ' '} {s['stationName'][:40]}")

def embed(txt):
    idx = HERE / "index.html"; html = idx.read_text(encoding="utf-8")
    block = '<script type="application/json" id="dataStore">\n' + txt.replace("</", "<\\/") + "\n</script>\n"
    html = re.sub(r'<script type="application/json" id="dataStore">.*?</script>\n?', "", html, flags=re.S)
    i = html.index("\n<script>\n"); idx.write_text(html[:i+1] + block + html[i+1:], encoding="utf-8")

if WRITE:
    txt = json.dumps(doc, ensure_ascii=False, indent=0)
    STATIONS_JSON.write_text(txt, encoding="utf-8"); embed(txt)
    print("WROTE stations.json + embedded into index.html")
else:
    print("(dry run — pass --write)")
